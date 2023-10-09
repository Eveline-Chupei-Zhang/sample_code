# -*- coding: utf-8 -*-
"""
web-scrapping: the box office data of US movies from IMDB

@author: Eveline Chupei Zhang
"""


import requests
import bs4
import re
import openpyxl
import os
from lxml import etree



def get_all_country(url,path):
    response = requests.get(url)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')
    books = soup.find("span", attrs={"class": "a-dropdown-container"})
    tags = books.select("span > select > option")
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    i=1
    for tag in tags:
        country=str(tag).split(">")[-2].split("</option")[0]
        ws[f'A{i}']=country
        print(country)
        href=tag.get("value")
        countrycode=href.split("area=")[1]
        ws[f'B{i}']=countrycode
        print(countrycode)
        i=i+1
    wb.save(path)


def get_box(soup,i):
    #response = requests.get(url)
    #soup = bs4.BeautifulSoup(response.text, 'html.parser')
    books = soup.find_all("div", attrs={"class": "a-section a-spacing-none mojo-performance-summary-table"})
    for book in books:
        tags=book.find_all("div", attrs={"class": "a-section a-spacing-none"})
        box_country=str(tags[0]).split("$")[-1].split("</span>")[0]
        ws_data[f'B{i}'] = box_country
        box_domestic=str(tags[1]).split("$")[-1].split("</span>")[0]
        ws_data[f'C{i}'] = box_domestic
        box_worldwide=str(tags[2]).split("$")[-1].split("</span>")[0]
        ws_data[f'D{i}'] = box_worldwide


def get_original_release_url(soup):
    #response = requests.get(url)
    #soup = bs4.BeautifulSoup(response.text, 'html.parser')
    book = soup.find_all("select", attrs={"name": "releasegroup-picker-navSelector"})
    if book:
        num=book[0].select("select > option")
        l=len(num)
        tag=str(book[0]).split('All')[-1].split('Original')[0]
        group_href=tag.split('="')[-1].split('">')[0]
        group_url="https://www.boxofficemojo.com"+group_href
        return group_url


def get_film_url(soup):
    #response = requests.get(url)
    #soup = bs4.BeautifulSoup(response.text, 'html.parser')
    text = soup.find("div", attrs={"class": "a-section a-spacing-none mojo-summary-values mojo-hidden-from-mobile"})
    items=text.select('div > div')
    tag=items[len(items)-1].select('div > span > a')
    href=tag[0].get("href")
    film_id=href.split("title/")[-1].split("/?ref")[0]
    ws_data[f'J{i}'] = film_id
    return(film_id)


 
def get_origin(film_id,i):
    url="https://www.imdb.com/title/"+film_id+"/?ref_=fn_al_tt_1"
    response = requests.get(url)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')
    #country of origin
    tag = soup.find("li", attrs={"data-testid": "title-details-origin"})
    origin=""
    if tag:
        ll = tag.select("li>div>ul>li>a")
        for l in ll:
            origin = str(origin)+str(l).split("rel=")[-1].split("</a>")[0]
    print(origin)
    ws_data[f'E{i}'] = origin
    #language
    tag3 = soup.find("li", attrs={"data-testid": "title-details-languages"})
    language=""
    if tag3:
        lans = tag3.select("li>div>ul>li>a")
        for lan in lans:
            language = str(language)+str(lan).split("rel=")[-1].split("</a>")[0]
    print(language)
    ws_data[f'F{i}'] = language
    #genre
    tag2 = soup.find("div", attrs={"data-testid": "genres"})
    genre=""
    if tag2:
        gg = tag2.select("div>a>span")
        for g in gg:
            genre = str(genre)+str(g).split("presentation")[-1].split("</span>")[0]
    print(genre)
    ws_data[f'G{i}'] = genre
    #rating
    tag3 = soup.find("div", attrs={"data-testid": "hero-rating-bar__aggregate-rating__score"})
    rating=""
    if tag3:
        rr= tag3.find("span", attrs={"class": "sc-7ab21ed2-1 jGRxWM"})
        rating=str(rr).split('">')[-1].split('</')[0]
    print(rating)
    ws_data[f'I{i}'] = rating



def get_movie_box(url):
    XLSX_PATH='/Users/jiaxin/Desktop/movie box/'+code+str(year)+"-"+str(i)+'.xlsx'
    wb_movie = openpyxl.Workbook()
    ws_movie = wb_movie['Sheet']
    response = requests.get(url)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')
    books = soup.find_all("table", attrs={"class": "a-bordered a-horizontal-stripes mojo-table releases-by-region"})
    k=1
    for book in books:
        tags=book.select("table > tr")
        for tag in tags:
            t=tag.select("tr > td")
            if t:
                ws_movie[f'A{k}'] = head
                text1=str(t[0]).split('">')[-1].split('</')[0]
                ws_movie[f'B{k}'] = text1
                text2=str(t[1]).split('">')[-1].split('</')[0]
                ws_movie[f'C{k}'] = text2
                text3=str(t[2]).split('">')[-1].split('</')[0]
                ws_movie[f'D{k}'] = text3
                text4=str(t[3]).split('">')[-1].split('</')[0]
                ws_movie[f'E{k}'] = text4   
                k=k+1
    wb_movie.save(XLSX_PATH) 


        
if __name__ == "__main__": 
    url="https://www.boxofficemojo.com/intl/?ref_=bo_nb_yld_tab"
    path="country.xlsx"
    #get_all_country(url,path)
    wb = openpyxl.load_workbook(path)
    ws=wb['Sheet']
    #code="US"
    #c_url="https://www.boxofficemojo.com/year/2018/"
    for year in range(2000,2020):
        for k in range(1,102):
            if ws[f'B{k}']:
                code=ws[f'B{k}'].value
                print(code)
                XLSX_PATH="/box/"+code+"-"+str(year)+".xlsx"
                c_url="https://www.boxofficemojo.com/year/"+str(year)+"/?area="+code+"&grossesOption=totalGrosses"
                wb_data = openpyxl.Workbook()
                ws_data=wb_data['Sheet']
                i=1
                response = requests.get(c_url)
                soup = bs4.BeautifulSoup(response.text, 'html.parser')
                books = soup.find_all("td", attrs={"class": "a-text-left mojo-field-type-release mojo-cell-wide"})
                if books:
                    for book in books:
                        item=book.select('td >a')
                        tag=item[0]
                        href=tag.get("href")
                        f_url="https://www.boxofficemojo.com/"+href
                        head=str(tag).split(">")[-2].split("</a")[0]
                        print(head)
                        ws_data[f'A{i}'] = head
                        response1 = requests.get(f_url)
                        soup1 = bs4.BeautifulSoup(response1.text, 'html.parser')
                        get_box(soup1,i)
                        group_url=get_original_release_url(soup1)
                        film_id=get_film_url(soup1)    
                        get_origin(film_id,i)
                        ws_data[f'H{i}'] = code
                        wb_data.save(XLSX_PATH)
                        if group_url:
                            get_movie_box(group_url)
                        i=i+1
    
    
    
    
    
